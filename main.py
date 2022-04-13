import openpyxl

#criar uma planilha
book = openpyxl.Workbook()
#visualizar paginas existentes
print(book.sheetnames)
#criar uma pagina
book.create_sheet('Frutas')
#selecionar uma pagina
frutas_pages = book['Frutas']
#conteudo das celulas
frutas_pages.append(['banana','5','R$3,90'])
frutas_pages.append(['laranja','14','R$2,40'])
frutas_pages.append(['goiaba','20','R$1,80'])
frutas_pages.append(['maçã','8','R$1,40'])
#salvar a planilha
book.save('Planilha de compras.xlsx')